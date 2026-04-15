"""Enhanced file tools — PDF reading, image OCR, Excel/CSV processing.

Optional dependencies:
- PDF: pip install pymupdf  (or: pip install cheetahclaws[files])
- OCR: pip install pytesseract Pillow  (+ system tesseract)
- Excel: pip install openpyxl  (or: pip install cheetahclaws[files])
"""
from __future__ import annotations

import csv
import io
from pathlib import Path

from tool_registry import ToolDef, register_tool


def _read_pdf(params: dict, config: dict) -> str:
    """Read text content from a PDF file."""
    try:
        import fitz  # pymupdf
    except ImportError:
        return (
            "PDF reading requires pymupdf. Install with:\n"
            "  pip install pymupdf\n"
            "Or: pip install cheetahclaws[files]"
        )

    file_path = params["file_path"]
    pages = params.get("pages")  # e.g. "1-5" or "3" or None (all)
    p = Path(file_path)

    if not p.exists():
        return f"Error: file not found: {file_path}"
    if not p.suffix.lower() == ".pdf":
        return f"Error: not a PDF file: {file_path}"

    try:
        doc = fitz.open(str(p))
        total = len(doc)

        # Parse page range
        if pages:
            page_list = _parse_page_range(pages, total)
        else:
            page_list = list(range(min(total, 50)))  # default: first 50 pages

        text_parts = []
        for i in page_list:
            if 0 <= i < total:
                page = doc[i]
                text = page.get_text()
                if text.strip():
                    text_parts.append(f"--- Page {i+1} ---\n{text.strip()}")

        doc.close()

        if not text_parts:
            return f"PDF has {total} pages but no extractable text (may be scanned/image-only)."

        header = f"PDF: {p.name} ({total} pages, showing {len(text_parts)})\n\n"
        content = "\n\n".join(text_parts)

        if len(content) > 50000:
            content = content[:50000] + f"\n\n[... truncated, {len(content)-50000} chars remaining ...]"

        return header + content

    except Exception as e:
        return f"Error reading PDF: {type(e).__name__}: {e}"


def _read_image(params: dict, config: dict) -> str:
    """Extract text from an image using OCR."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return (
            "OCR requires pytesseract and Pillow. Install with:\n"
            "  pip install pytesseract Pillow\n"
            "Also install Tesseract OCR engine:\n"
            "  macOS: brew install tesseract\n"
            "  Ubuntu: sudo apt install tesseract-ocr\n"
            "  Windows: https://github.com/UB-Mannheim/tesseract/wiki"
        )

    file_path = params["file_path"]
    lang = params.get("language", "eng")
    p = Path(file_path)

    if not p.exists():
        return f"Error: file not found: {file_path}"

    try:
        img = Image.open(str(p))
        text = pytesseract.image_to_string(img, lang=lang)

        if not text.strip():
            return f"No text detected in image: {p.name}"

        return f"OCR result from {p.name} ({img.size[0]}x{img.size[1]}, lang={lang}):\n\n{text.strip()}"

    except Exception as e:
        return f"OCR error: {type(e).__name__}: {e}"


def _read_excel(params: dict, config: dict) -> str:
    """Read data from Excel (.xlsx) or CSV files."""
    file_path = params["file_path"]
    sheet = params.get("sheet")
    max_rows = min(params.get("max_rows", 100), 500)
    p = Path(file_path)

    if not p.exists():
        return f"Error: file not found: {file_path}"

    ext = p.suffix.lower()

    try:
        if ext == ".csv":
            return _read_csv_file(p, max_rows)
        elif ext in (".xlsx", ".xls"):
            return _read_xlsx_file(p, sheet, max_rows)
        elif ext == ".tsv":
            return _read_csv_file(p, max_rows, delimiter="\t")
        else:
            return f"Unsupported format: {ext}. Supported: .csv, .tsv, .xlsx, .xls"

    except Exception as e:
        return f"Error reading {ext} file: {type(e).__name__}: {e}"


def _read_csv_file(path: Path, max_rows: int, delimiter: str = ",") -> str:
    """Read a CSV/TSV file."""
    with open(path, encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row)

    if not rows:
        return f"Empty file: {path.name}"

    # Format as table
    return _format_table(rows, path.name, total_hint=f"showing {len(rows)} rows")


def _read_xlsx_file(path: Path, sheet: str | None, max_rows: int) -> str:
    """Read an Excel file."""
    try:
        import openpyxl
    except ImportError:
        return (
            "Excel reading requires openpyxl. Install with:\n"
            "  pip install openpyxl\n"
            "Or: pip install cheetahclaws[files]"
        )

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if sheet and sheet in sheet_names:
        ws = wb[sheet]
    elif sheet and sheet.isdigit():
        idx = int(sheet) - 1
        if 0 <= idx < len(sheet_names):
            ws = wb[sheet_names[idx]]
        else:
            wb.close()
            return f"Sheet index {sheet} out of range. Available: {', '.join(sheet_names)}"
    else:
        ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([str(c) if c is not None else "" for c in row])

    wb.close()

    if not rows:
        return f"Empty sheet: {ws.title}"

    header = f"Excel: {path.name}, sheet: {ws.title}"
    if len(sheet_names) > 1:
        header += f" (sheets: {', '.join(sheet_names)})"

    return _format_table(rows, header, total_hint=f"showing {len(rows)} rows")


def _format_table(rows: list[list], title: str, total_hint: str = "") -> str:
    """Format rows as a readable text table."""
    if not rows:
        return "(empty)"

    # Calculate column widths (cap at 30 chars per col)
    n_cols = max(len(r) for r in rows)
    widths = [0] * n_cols
    for row in rows[:20]:  # sample first 20 rows for width
        for j, cell in enumerate(row):
            if j < n_cols:
                widths[j] = min(max(widths[j], len(str(cell))), 30)

    lines = [f"{title} ({total_hint})\n"]

    for i, row in enumerate(rows):
        cells = []
        for j in range(n_cols):
            val = row[j] if j < len(row) else ""
            cells.append(str(val)[:30].ljust(widths[j]))
        lines.append(" | ".join(cells))
        if i == 0:
            lines.append("-+-".join("-" * w for w in widths))

    return "\n".join(lines)


def _parse_page_range(spec: str, total: int) -> list[int]:
    """Parse page range like '1-5', '3', '1,3,5-8'."""
    pages = []
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            start = max(int(a) - 1, 0)
            end = min(int(b), total)
            pages.extend(range(start, end))
        elif part.isdigit():
            pages.append(int(part) - 1)
    return sorted(set(pages))


# ── Register ─────────────────────────────────────────────────────────────

register_tool(ToolDef(
    name="ReadPDF",
    schema={
        "name": "ReadPDF",
        "description": (
            "Read text content from a PDF file. Extracts text from specified pages. "
            "For scanned/image-only PDFs, use ReadImage with OCR instead."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the PDF file",
                },
                "pages": {
                    "type": "string",
                    "description": "Page range: '1-5', '3', '1,3,5-8'. Default: first 50 pages.",
                },
            },
            "required": ["file_path"],
        },
    },
    func=_read_pdf,
    read_only=True,
    concurrent_safe=True,
))

register_tool(ToolDef(
    name="ReadImage",
    schema={
        "name": "ReadImage",
        "description": (
            "Extract text from an image using OCR (Tesseract). "
            "Supports PNG, JPG, TIFF, BMP. Useful for scanned documents, screenshots, "
            "and image-only PDFs (convert to image first)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the image file",
                },
                "language": {
                    "type": "string",
                    "description": "OCR language code (default: eng). Examples: chi_sim (Chinese), jpn (Japanese), deu (German)",
                    "default": "eng",
                },
            },
            "required": ["file_path"],
        },
    },
    func=_read_image,
    read_only=True,
    concurrent_safe=True,
))

register_tool(ToolDef(
    name="ReadSpreadsheet",
    schema={
        "name": "ReadSpreadsheet",
        "description": (
            "Read data from Excel (.xlsx/.xls) or CSV/TSV files. "
            "Returns formatted table with column alignment. "
            "For Excel files with multiple sheets, specify the sheet name or number."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the spreadsheet file (.xlsx, .xls, .csv, .tsv)",
                },
                "sheet": {
                    "type": "string",
                    "description": "Sheet name or number (Excel only, default: active sheet)",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "Max rows to return (default: 100, max: 500)",
                    "default": 100,
                },
            },
            "required": ["file_path"],
        },
    },
    func=_read_excel,
    read_only=True,
    concurrent_safe=True,
))
